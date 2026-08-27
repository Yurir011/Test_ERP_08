import datetime
import json
import os
import re
import sys
import winsound

import openpyxl
import win32com.client
from PyQt5 import uic
from PyQt5.QtCore import QDate, Qt
from PyQt5.QtGui import QFontMetricsF, QIcon, QPainter
from PyQt5.QtPrintSupport import QPrinter
from PyQt5.QtWidgets import (
    QApplication,
    QFileDialog,
    QLabel,
    QMainWindow,
    QMessageBox,
    QScrollArea,
    QTextBrowser,
    QTextEdit,
)

### ㄲ뜸ㄲ ###
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UI_DIR = os.path.join(BASE_DIR, "UI")

HOME_UI = os.path.join(UI_DIR, "Home.ui")
CLICK_SOUND = os.path.join(BASE_DIR, "sounds", "click.wav")

QUOTATION_UI = os.path.join(UI_DIR, "Quotation.ui")
QUOTATION2_UI = os.path.join(UI_DIR, "Quotation2.ui")
QUOTATION_XLSX = os.path.join(BASE_DIR, "BN_견적서_양식.xlsx")
QUOTATION_SAVE_DIR = os.path.join(BASE_DIR, "List", "Quotation_List")
QUOTATION_LIST_XLSX = os.path.join(QUOTATION_SAVE_DIR, "List", "견적서List.xlsx")
QUOTATION_SEQ_FILE = os.path.join(BASE_DIR, "quotation_seq.json")

DELIVERY_NOTE_UI = os.path.join(UI_DIR, "DeliveryNote.ui")
DELIVERY_NOTE2_UI = os.path.join(UI_DIR, "DeliveryNote2.ui")
DELIVERY_NOTE_XLSX = os.path.join(BASE_DIR, "BN_거래명세서_양식.xlsx")
DELIVERY_NOTE_SAVE_DIR = os.path.join(BASE_DIR, "List", "DeliveryNott_List")
DELIVERY_NOTE_LIST_XLSX = os.path.join(DELIVERY_NOTE_SAVE_DIR, "List", "거래명세서List.xlsx")
DELIVERY_NOTE_SEQ_FILE = os.path.join(BASE_DIR, "delivery_note_seq.json")

DOC_UI = os.path.join(UI_DIR, "Doc.ui")
DOC2_UI = os.path.join(UI_DIR, "Doc2.ui")
DOC_XLSX = os.path.join(BASE_DIR, "BN_공문서_양식.xlsx")
DOC_SAVE_DIR = os.path.join(BASE_DIR, "List", "OfficialDoc_List")
DOC_LIST_XLSX = os.path.join(DOC_SAVE_DIR, "List", "공문서List.xlsx")
DOC_SEQ_FILE = os.path.join(BASE_DIR, "doc_seq.json")
DOC_APPROVE_ICON = os.path.join(BASE_DIR, "icon-approve.png")
DOC_REJECT_ICON = os.path.join(BASE_DIR, "icon-reject.png")

XL_TYPE_PDF = 0

# Each item row on Quotation2.ui as (item, description, quantity, unit_price, amount_display)
# widget names. Rows correspond, in order, to the quotation template's item rows 16-23.
ITEM_ROWS = (
    ("Item1_textEdit", "Description1_textEdit", "Qty1_textEdit", "UnitPrice1_textEdit", "Amount1_textEdit"),
    ("Item2_textEdit", "Description2_textEdit", "Qty2_textEdit", "UnitPrice2_textEdit", "Amount2_textEdit"),
    ("Item3_textEdit", "Description3_textEdit", "Qty3_textEdit", "UnitPrice3_textEdit", "Amount3_textEdit"),
    ("Item4_textEdit", "Description4_textEdit", "Qty4_textEdit", "UnitPrice4_textEdit", "Amount4_textEdit"),
    ("Item5_textEdit", "Description5_textEdit", "Qty5_textEdit", "UnitPrice5_textEdit", "Amount5_textEdit"),
    ("Item6_textEdit", "Description6_textEdit", "Qty6_textEdit", "UnitPrice6_textEdit", "Amount6_textEdit"),
    ("Item7_textEdit", "Description7_textEdit", "Qty7_textEdit", "UnitPrice7_textEdit", "Amount7_textEdit"),
    ("Item8_textEdit", "Description8_textEdit", "Qty8_textEdit", "UnitPrice8_textEdit", "Amount8_textEdit"),
)

# (excel cell, Quotation2.ui widget name) for the single-value header/customer fields.
HEADER_CELL_MAP = (
    ("G2", "textEdit_6"),  # date
    ("G3", "textEdit_7"),  # quotation number
    ("A11", "ClientName_textEdit"),  # customer name
    ("A12", "ClientAddress_textEdit"),  # customer address
)
REMARKS_CELL = "F10"  # textEdit_2

# Same item-row layout on DeliveryNote2.ui, just under its own (slightly differently
# named) widgets.
DELIVERY_ITEM_ROWS = (
    ("Item1_textEdit", "Description1_textEdit", "Qty1_textEdit", "UnitPrice1_textEdit", "Amount1_textEdit"),
    ("Item2_textEdit", "Description2_textEdit", "Qty2_textEdit", "UnitPrice2_textEdit", "Amount2_textEdit"),
    ("Item3_textEdit", "Description3_textEdit", "Qty3_textEdit", "UnitPrice3_textEdit", "Amount3_textEdit"),
    ("Item4_textEdit", "Description4_textEdit", "Qty4_textEdit", "UnitPrice4_textEdit", "Amount4_textEdit"),
    ("Item5_textEdit", "Description5_textEdit", "Qty5_textEdit", "UnitPrice5_textEdit", "Amount5_textEdit"),
    ("Item6_textEdit", "Description6_textEdit", "Qty6_textEdit", "UnitPrice6_textEdit", "Amount6_textEdit"),
    ("Item7_textEdit", "Description7_textEdit", "Qty7_textEdit", "UnitPrice7_textEdit", "Amount7_textEdit"),
    ("Item8_textEdit", "Description8_textEdit", "Qty8_textEdit", "UnitPrice8_textEdit", "Amount8_textEdit"),
)
DELIVERY_HEADER_CELL_MAP = (
    ("G2", "Date_textEdit"),  # date
    ("G3", "textEdit_7"),  # document number
    ("A11", "ClientNametextEdit"),  # customer name
    ("A12", "ClientAdd_textEdit"),  # customer address
)
DELIVERY_REMARKS_CELL = "F10"  # textEdit_2

MAX_ITEM_ROWS = 8  # item rows supported by the Quotation2/DeliveryNote2 templates

# Rounded, bordered look for the compact forms' input fields (Quotation.ui / DeliveryNote.ui /
# Doc.ui), replacing the default flat QTextBrowser frame. Paired with turning the vertical
# scrollbar off wherever this is applied, so a stray scrollbar thumb never appears either.
COMPACT_INPUT_STYLESHEET = (
    "QTextBrowser {"
    " background: white;"
    " border: 1px solid #c3c3c3;"
    " border-radius: 8px;"
    " padding: 0 10px;"
    "}"
)

# --- Quotation.ui / DeliveryNote.ui (compact entry forms): identical item-row grid and
# widget names in both files, only differing in the button cluster that slides down
# when rows don't fit, and in which full-page window/number-prefix Go hands off to.
COMPACT_ITEM_ROW_COLUMNS = (
    ("item", 80, 201),
    ("description", 290, 431),
    ("quantity", 730, 91),
    ("unit_price", 830, 201),
)
COMPACT_ITEM_ROW_TOP = 400
COMPACT_ITEM_ROW_HEIGHT = 46
COMPACT_ITEM_ROW_SPACING = 46
COMPACT_ITEM_ROW_GAP = 20
QUOTATION_BOTTOM_CLUSTER_WIDGET_NAMES = (
    "Plus_pushButton",
    "Minus_pushButton",
    "Go_pushButton",
    "Back_pushButton",
    "pushButton_3",
    "label_9",
    "label_11",
    "label_12",
)
DELIVERY_NOTE_BOTTOM_CLUSTER_WIDGET_NAMES = QUOTATION_BOTTOM_CLUSTER_WIDGET_NAMES + ("Go_pushButton_2", "label_13")

# --- Doc2.ui (공문서 full-page document): single-value header fields, a free-form
# body split one line per template row, and a 4-stage approval strip.
DOC_HEADER_CELL_MAP = (
    ("C6", "Number_textEdit"),  # 문서번호
    ("C7", "Date_textEdit"),  # 시행일자
    ("C8", "textEdit_3"),  # 경유
    ("C9", "textEdit_4"),  # 수신
    ("C10", "Title_textEdit"),  # 제목
)
DOC_PERSON_CELL = "A3"  # letterhead contact person - cell already reads "담당자     <name>"
DOC_PERSON_LABEL = "담당자"
DOC_FOOTER_PERSON_CELL = "A40"
DOC_FOOTER_COOP_CELL = "A41"
DOC_BODY_ROWS = tuple(r for r in range(15, 39) if r != 36)  # one MainBody_textEdit line
# per row, A15..A38 - skipping row 36, which holds the fixed company signature text.

# (approval button widget name, excel cell for that stage's stamp/status, role label)
DOC_APPROVAL_STAGES = (
    ("pushButton_4", "H7", "담당자"),
    ("pushButton", "I7", "관리자"),
    ("pushButton_2", "J7", "부서장"),
    ("pushButton_3", "K7", "대표이사"),
)

# Newer Qt Designer saves enums in Qt6's scoped form (e.g. "Qt::Orientation::Horizontal"),
# which PyQt5's uic parser (Qt5) does not understand. Normalize to the old flat form
# before parsing so re-saving in Designer can't break loading.
_SCOPED_ENUM_RE = re.compile(r"\b(\w+)::\w+::(\w+)\b")


def load_ui(ui_path, baseinstance):
    with open(ui_path, "r", encoding="utf-8") as f:
        content = f.read()
    normalized = _SCOPED_ENUM_RE.sub(r"\1::\2", content)
    if normalized == content:
        uic.loadUi(ui_path, baseinstance)
        return

    tmp_path = ui_path + ".normalized.tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        f.write(normalized)
    try:
        uic.loadUi(tmp_path, baseinstance)
    finally:
        os.remove(tmp_path)


def center_text_vertically(text_edit):
    """Vertically center text_edit's content, and keep it centered as the text
    changes (typing, prefill, live amount recalculation, 문서참조 loads, etc.)."""

    doc = text_edit.document()
    # The document's own top/bottom margin sits inside its reported size, so splitting
    # that size in half as a single top-only viewport margin double-counts it on top
    # and leaves it (plus any leftover viewport space) to silently pad the bottom
    # instead - text ends up sitting visibly above center. Zero it out and center
    # against the frame explicitly, in equal top/bottom viewport margins, instead.
    doc.setDocumentMargin(0)

    def recenter():
        no_wrap = text_edit.lineWrapMode() == QTextEdit.NoWrap
        doc.setTextWidth(-1 if no_wrap else text_edit.width())
        if no_wrap:
            # QTextDocument's own line-box height bakes in each font's line-spacing/leading,
            # which for some (esp. CJK) fonts runs far taller than the glyphs actually drawn -
            # centering against that box clamps the margin to 0 and pins single-line text to
            # the top with a large empty gap below. QFontMetrics' ascent+descent tracks the
            # visible ink instead, so single-line fields center the way a QLineEdit would.
            content_height = QFontMetricsF(text_edit.currentFont()).height()
        else:
            content_height = doc.size().height()
        # contentsRect() (not frameWidth()) accounts for the vertical inset correctly even
        # when a stylesheet gives the box asymmetric horizontal/vertical border+padding -
        # frameWidth() collapses that into one number and gets the vertical space wrong.
        available = text_edit.contentsRect().height()
        margin = int(max(0, (available - content_height) / 2))
        text_edit.setViewportMargins(0, margin, 0, margin)

    recenter()
    if not text_edit.property("_center_text_live"):
        text_edit.setProperty("_center_text_live", True)
        text_edit.textChanged.connect(recenter)


def next_document_number(prefix, seq_file):
    date_str = QDate.currentDate().toString("yyyyMMdd")
    seq = 0
    if os.path.exists(seq_file):
        with open(seq_file, "r", encoding="utf-8") as f:
            state = json.load(f)
        if state.get("date") == date_str:
            seq = (state.get("seq", -1) + 1) % 1000

    with open(seq_file, "w", encoding="utf-8") as f:
        json.dump({"date": date_str, "seq": seq}, f)

    return f"{prefix}-{date_str}-{seq:03d}"


def _advance_on_enter(widget, advance):
    original_key_press = widget.keyPressEvent

    def key_press(event):
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            advance()
            event.accept()
            return
        original_key_press(event)

    widget.keyPressEvent = key_press


def _as_number(text):
    try:
        return int(text)
    except ValueError:
        pass
    try:
        return float(text)
    except ValueError:
        return text


def export_pdf(xlsx_path, pdf_path):
    excel = win32com.client.DispatchEx("Excel.Application")
    try:
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(xlsx_path)
        try:
            workbook.ExportAsFixedFormat(XL_TYPE_PDF, pdf_path)
        finally:
            workbook.Close(False)
    finally:
        excel.Quit()


def export_widget_pdf(widget, pdf_path):
    """Render exactly what's on screen (the document page as filled in) to a PDF,
    independent of the Excel-generated one."""
    pixmap = widget.grab()

    printer = QPrinter(QPrinter.HighResolution)
    printer.setOutputFormat(QPrinter.PdfFormat)
    printer.setOutputFileName(pdf_path)
    printer.setPageSize(QPrinter.A4)
    printer.setPageMargins(0, 0, 0, 0, QPrinter.Millimeter)

    painter = QPainter(printer)
    try:
        page_rect = printer.pageRect()
        scaled = pixmap.scaled(page_rect.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
        x = page_rect.x() + (page_rect.width() - scaled.width()) // 2
        y = page_rect.y() + (page_rect.height() - scaled.height()) // 2
        painter.drawPixmap(x, y, scaled)
    finally:
        painter.end()


def export_document_files(workbook, doc_no, save_dir, source_widget):
    """Save `workbook` as {doc_no}.xlsx into save_dir, export it to {doc_no}.pdf via
    Excel, and grab a {doc_no}_screen.pdf of source_widget. Returns the xlsx path."""
    os.makedirs(save_dir, exist_ok=True)

    xlsx_path = os.path.join(save_dir, f"{doc_no}.xlsx")
    workbook.save(xlsx_path)

    pdf_path = os.path.join(save_dir, f"{doc_no}.pdf")
    export_pdf(xlsx_path, pdf_path)

    screen_pdf_path = os.path.join(save_dir, f"{doc_no}_screen.pdf")
    export_widget_pdf(source_widget, screen_pdf_path)

    return xlsx_path


def append_to_list_xlsx(list_xlsx_path, date_text, name, item, description, quantity, unit_price):
    workbook = openpyxl.load_workbook(list_xlsx_path)
    sheet = workbook.active

    row = 3
    while sheet[f"A{row}"].value is not None:
        row += 1

    sheet[f"A{row}"] = row - 2
    sheet[f"B{row}"] = date_text
    sheet[f"C{row}"] = name
    sheet[f"D{row}"] = item
    sheet[f"E{row}"] = description
    sheet[f"J{row}"] = quantity
    sheet[f"K{row}"] = unit_price
    sheet[f"L{row}"] = f"=J{row}*K{row}"

    workbook.save(list_xlsx_path)


def append_to_doc_list(list_xlsx_path, date_text, manager, number, routing, recipient, title):
    workbook = openpyxl.load_workbook(list_xlsx_path)
    sheet = workbook.active

    row = 3
    while sheet[f"B{row}"].value is not None:
        row += 1

    sheet[f"B{row}"] = date_text
    sheet[f"C{row}"] = manager
    sheet[f"D{row}"] = number
    sheet[f"E{row}"] = routing
    sheet[f"F{row}"] = recipient
    sheet[f"G{row}"] = title

    workbook.save(list_xlsx_path)


# --- "문서 참조" (load a previously saved xlsx/pdf back into the compact entry form).
# Both parsers read the same cell/text layout save_filled_document() writes, so a
# hand-edited or differently formatted file can throw them off - this is a best-effort
# convenience, not a guaranteed-correct import.

def parse_saved_xlsx(xlsx_path):
    workbook = openpyxl.load_workbook(xlsx_path)
    sheet = workbook.active

    def cell_text(coord):
        value = sheet[coord].value
        if value is None:
            return ""
        if isinstance(value, (datetime.date, datetime.datetime)):
            return value.strftime("%Y-%m-%d")
        return str(value).strip()

    rows = []
    for row_num in range(16, 24):
        item = cell_text(f"A{row_num}")
        desc = cell_text(f"B{row_num}")
        qty = cell_text(f"G{row_num}")
        price = cell_text(f"H{row_num}")
        if item or desc or qty or price:
            rows.append((item, desc, qty, price))

    return {
        "date": cell_text("G2"),
        "number": cell_text("G3"),
        "name": cell_text("A11"),
        "address": cell_text("A12"),
        "rows": rows,
    }


def _cluster_pdf_words(words, tolerance=2.0):
    """Group words (from pdfplumber's extract_words()) into lines by their 'top'
    coordinate: top-to-bottom, each line sorted left-to-right by x0."""
    lines = []
    for word in sorted(words, key=lambda w: w["top"]):
        for line in lines:
            if abs(line[0]["top"] - word["top"]) <= tolerance:
                line.append(word)
                break
        else:
            lines.append([word])
    for line in lines:
        line.sort(key=lambda w: w["x0"])
    lines.sort(key=lambda line: line[0]["top"])
    return lines


def _cluster_pdf_words_into_lines(words, tolerance=2.0):
    return [" ".join(w["text"] for w in line) for line in _cluster_pdf_words(words, tolerance)]


def parse_saved_pdf(pdf_path):
    import pdfplumber  # imported lazily so a missing dependency only breaks this feature

    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
        table = page.extract_table() or []

    full_text = " ".join(w["text"] for w in words)
    date_match = re.search(r"\d{4}-\d{2}-\d{2}", full_text)
    number_match = re.search(r"BN-[A-Z]+-\d{8}-\d{3}", full_text)

    # Customer name/address sit between the "Email : ..." line (end of our fixed
    # company info block) and the "아래와 같이 ~합니다." line (start of the item
    # table section) on both templates - the first line in between is a one-word
    # section label ("고객"/"납품") to skip, then name, then address.
    email_words = [w for w in words if w["text"] == "Email"]
    start_words = [w for w in words if w["text"].startswith("아래와")]
    name, address = "", ""
    if email_words and start_words:
        top_bound = email_words[0]["bottom"]
        bottom_bound = start_words[0]["top"]
        between = [w for w in words if top_bound < w["top"] < bottom_bound]
        content_lines = _cluster_pdf_words_into_lines(between)[1:]
        if content_lines:
            name = content_lines[0]
        if len(content_lines) > 1:
            address = content_lines[1]

    rows = []
    header_idx = next((i for i, row in enumerate(table) if row and row[0] == "ITEM #"), None)
    if header_idx is not None:
        for row in table[header_idx + 1:]:
            if not row or row[0] == "TOTAL":
                break
            item = (row[0] or "").strip()
            desc_candidates = [c for c in row[1:6] if c]
            desc = desc_candidates[0].strip() if desc_candidates else ""
            qty = (row[6] or "").strip()
            price = (row[7] or "").strip()
            if item or desc or qty or price:
                rows.append((item, desc, qty, price))
            if len(rows) >= MAX_ITEM_ROWS:
                break

    return {
        "date": date_match.group(0) if date_match else "",
        "number": number_match.group(0) if number_match else "",
        "name": name,
        "address": address,
        "rows": rows,
    }


def parse_saved_doc_xlsx(xlsx_path):
    workbook = openpyxl.load_workbook(xlsx_path)
    sheet = workbook.active

    def cell_text(coord):
        value = sheet[coord].value
        if value is None:
            return ""
        if isinstance(value, (datetime.date, datetime.datetime)):
            return value.strftime("%Y-%m-%d")
        return str(value).strip()

    person = re.sub(rf"^{DOC_PERSON_LABEL}\s*", "", cell_text(DOC_PERSON_CELL)).strip()

    body_lines = [cell_text(f"A{row}") for row in DOC_BODY_ROWS]
    while body_lines and not body_lines[-1]:
        body_lines.pop()

    return {
        "number": cell_text("C6"),
        "date": cell_text("C7"),
        "person": person,
        "routing": cell_text("C8"),
        "recipient": cell_text("C9"),
        "title": cell_text("C10"),
        "body": "\n".join(body_lines),
    }


def parse_saved_doc_pdf(pdf_path):
    import pdfplumber  # imported lazily so a missing dependency only breaks this feature

    with pdfplumber.open(pdf_path) as pdf:
        words = pdf.pages[0].extract_words(use_text_flow=False, keep_blank_chars=False)

    lines = _cluster_pdf_words(words)
    full_text = " ".join(w["text"] for w in words)
    number_match = re.search(r"BN-[A-Z]+-\d{8}-\d{3}", full_text)

    # 문서번호/시행일자/경유/수신/제목 values all start at the same fixed x-column right
    # after their ":" label (~x0 80-90), stacked top-to-bottom in that order - a
    # reliable anchor regardless of how long each value is. Limited to top<250 so
    # body-text words that happen to land near that column don't get mistaken for them.
    field_matches = []
    for line in lines:
        if line[0]["top"] >= 250:
            continue
        value_start = next((w for w in line if 78 <= w["x0"] <= 90), None)
        if value_start is not None:
            value_words = [w for w in line if w["x0"] >= value_start["x0"] - 1 and w["x0"] < 250]
            field_matches.append((line[0]["top"], " ".join(w["text"] for w in value_words)))

    values = [v for _, v in field_matches[:5]] + [""] * 5
    number, date, routing, recipient, title = values[:5]
    if number_match:
        number = number_match.group(0)

    # Body runs from just below the 제목 line to the company signature line ("주식회사 ...").
    title_top = field_matches[4][0] if len(field_matches) >= 5 else None
    sig_line = next((line for line in lines if any(w["text"] == "주식회사" for w in line)), None)
    sig_top = sig_line[0]["top"] if sig_line else None
    body_lines = []
    if title_top is not None:
        for line in lines:
            top = line[0]["top"]
            if title_top + 5 < top < (sig_top if sig_top is not None else 1e9):
                body_lines.append(" ".join(w["text"] for w in line))
    body = "\n".join(body_lines)

    # Letterhead contact person ("담당자 <name>", above the field block, top<150).
    person = ""
    for line in lines:
        if line[0]["top"] < 150 and any(w["text"] == DOC_PERSON_LABEL for w in line):
            rest = [w for w in line if w["text"] != DOC_PERSON_LABEL and w["x0"] < 200]
            person = " ".join(w["text"] for w in rest)
            break

    return {
        "number": number,
        "date": date,
        "person": person,
        "routing": routing,
        "recipient": recipient,
        "title": title,
        "body": body,
    }


def make_window_scrollable(window):
    """Wrap window's centralwidget in a QScrollArea and cap the window to the screen's
    available size. The full-page templates (Doc2.ui / Quotation2.ui / DeliveryNote2.ui)
    are taller than most displays (up to ~1519px), so opened at their natural fixed size
    the bottom controls (Save/Print/...) land off-screen with no way to reach them."""
    # Right after load_ui, the window already has its .ui-declared size, but the
    # centralwidget hasn't been laid out to fill it yet (that happens lazily) - so read
    # the size from the window, not the not-yet-sized centralwidget, and stamp it on
    # explicitly since setWidgetResizable(False) below needs a real size to work with.
    full_size = window.size()
    content = window.centralWidget()
    content.resize(full_size)

    scroll_area = QScrollArea()
    scroll_area.setWidget(content)
    scroll_area.setWidgetResizable(False)
    scroll_area.setFrameShape(QScrollArea.NoFrame)
    window.setCentralWidget(scroll_area)

    available = QApplication.primaryScreen().availableGeometry()
    scrollbar_allowance = scroll_area.verticalScrollBar().sizeHint().width()
    width = min(full_size.width() + scrollbar_allowance, available.width())
    height = min(full_size.height(), available.height() - 60)
    window.setFixedSize(width, height)


# --- Shared logic for the "full-page" document windows (Quotation2.ui / DeliveryNote2.ui):
# fixed 8 item rows, live QTY*UnitPrice -> Amount, and a Save that fills the Excel
# template, exports it (+ a screen-capture PDF), and logs each item row to the running list.

def style_item_form(window, item_rows, extra_single_line_widgets):
    for text_edit in window.centralwidget.findChildren(QTextEdit):
        center_text_vertically(text_edit)

    single_line_widgets = list(extra_single_line_widgets)
    for row in item_rows:
        single_line_widgets.extend(getattr(window, name) for name in row)
    for widget in single_line_widgets:
        widget.setLineWrapMode(QTextEdit.NoWrap)
        widget.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

    for *_rest, amount_name in item_rows:
        getattr(window, amount_name).setReadOnly(True)


def connect_amount_calculations(window, item_rows):
    for _item_name, _desc_name, qty_name, price_name, amount_name in item_rows:
        qty_widget = getattr(window, qty_name)
        price_widget = getattr(window, price_name)
        amount_widget = getattr(window, amount_name)

        def recalc(_checked=False, qty_widget=qty_widget, price_widget=price_widget, amount_widget=amount_widget):
            qty = _as_number(qty_widget.toPlainText().strip())
            price = _as_number(price_widget.toPlainText().strip())
            if isinstance(qty, (int, float)) and isinstance(price, (int, float)):
                amount_widget.setPlainText(f"{qty * price:,.0f}")
            else:
                amount_widget.clear()

        qty_widget.textChanged.connect(recalc)
        price_widget.textChanged.connect(recalc)


def build_entry_chain(window, item_rows, leading_fields, submit_button):
    entry_fields = list(leading_fields)
    for item_name, desc_name, qty_name, price_name, _amount_name in item_rows:
        entry_fields.extend(getattr(window, n) for n in (item_name, desc_name, qty_name, price_name))

    for field in entry_fields:
        field.setTabChangesFocus(True)
    for current, following in zip(entry_fields, entry_fields[1:]):
        window.setTabOrder(current, following)
        _advance_on_enter(current, following.setFocus)
    window.setTabOrder(entry_fields[-1], submit_button)
    _advance_on_enter(entry_fields[-1], submit_button.click)
    return entry_fields


def save_filled_document(
    window,
    *,
    item_rows,
    header_cell_map,
    remarks_cell,
    remarks_widget_name,
    xlsx_template,
    save_dir,
    list_xlsx_path,
    name_widget_name,
    date_widget_name,
    number_widget_name,
):
    """Fill `xlsx_template` from `window`'s fields, export it + a screen-capture PDF into
    `save_dir`, and log each non-blank item row into `list_xlsx_path`. Returns the
    document number used."""
    rows_data = [
        (
            getattr(window, item_name).toPlainText().strip(),
            getattr(window, desc_name).toPlainText().strip(),
            getattr(window, qty_name).toPlainText().strip(),
            getattr(window, price_name).toPlainText().strip(),
        )
        for item_name, desc_name, qty_name, price_name, _amount_name in item_rows
    ]

    workbook = openpyxl.load_workbook(xlsx_template)
    sheet = workbook.active

    for cell, widget_name in header_cell_map:
        sheet[cell] = getattr(window, widget_name).toPlainText().strip()
    sheet[remarks_cell] = getattr(window, remarks_widget_name).toPlainText().strip()

    for offset, (item, desc, qty, price) in enumerate(rows_data):
        row = 16 + offset
        sheet[f"A{row}"] = item
        sheet[f"B{row}"] = desc
        sheet[f"G{row}"] = _as_number(qty) if qty else None
        sheet[f"H{row}"] = _as_number(price) if price else None

    doc_no = getattr(window, number_widget_name).toPlainText().strip()
    export_document_files(workbook, doc_no, save_dir, window.centralwidget)

    list_date = QDate.fromString(getattr(window, date_widget_name).toPlainText().strip(), "yyyy-MM-dd").toString("yyMMdd")
    name = getattr(window, name_widget_name).toPlainText().strip()
    for item, desc, qty, price in rows_data:
        if not item:
            continue
        append_to_list_xlsx(
            list_xlsx_path,
            list_date,
            name,
            item,
            desc,
            _as_number(qty) if qty else None,
            _as_number(price) if price else None,
        )

    return doc_no


class CompactItemFormWindow(QMainWindow):
    """Base for the compact entry forms (Quotation.ui / DeliveryNote.ui): a few header
    fields plus up to 8 dynamically add/remove-able item rows. Go hands everything off
    to the paired full-page window (set via `full_page_window_class`) for review/save."""

    ui_path = None
    window_title = ""
    number_prefix = None
    seq_file = None
    bottom_cluster_names = ()
    full_page_window_class = None

    def __init__(self):
        super().__init__()
        load_ui(self.ui_path, self)
        self.setFixedSize(self.size())
        self.setWindowTitle(self.window_title)

        # Date/number are auto-filled below and left read-only; customer name/address are
        # the only header fields the user actually types into.
        auto_filled = ("Date_textBrowser", "Number_textBrowse")
        for name in ("Date_textBrowser", "Number_textBrowse", "Name_textBrowse", "Address_textBrowse"):
            widget = getattr(self, name)
            widget.setReadOnly(name in auto_filled)
            widget.setLineWrapMode(QTextEdit.NoWrap)
            widget.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            widget.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            widget.setStyleSheet(COMPACT_INPUT_STYLESHEET)
            font = widget.font()
            font.setPointSize(12)
            widget.setFont(font)
            center_text_vertically(widget)

        self.Date_textBrowser.setText(QDate.currentDate().toString("yyyy-MM-dd"))
        self.Number_textBrowse.setText(next_document_number(self.number_prefix, self.seq_file))

        self._base_window_size = self.size()
        self._bottom_cluster_origin = {
            getattr(self, name): getattr(self, name).pos() for name in self.bottom_cluster_names
        }
        self._bottom_cluster_top = min(pos.y() for pos in self._bottom_cluster_origin.values())

        self.item_rows = [[
            self.Item_textBrowse,
            self.Description_textBrowse,
            self.Quantity_textBrowse,
            self.UnitPrice_textBrowse,
        ]]
        for widget in self.item_rows[0]:
            self._style_row_widget(widget)

        self._rebuild_entry_chain()
        self._update_bottom_layout()
        self.Name_textBrowse.setFocus()

        self._next_window = None
        self.Go_pushButton.clicked.connect(self.open_full_page)
        self.Back_pushButton.clicked.connect(self.close)
        self.Plus_pushButton.clicked.connect(self.add_item_row)
        self.Minus_pushButton.clicked.connect(self.remove_item_row)

    @staticmethod
    def _style_row_widget(widget):
        widget.setReadOnly(False)
        font = widget.font()
        font.setPointSize(12)
        widget.setFont(font)
        widget.setLineWrapMode(QTextEdit.NoWrap)
        widget.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        widget.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        widget.setStyleSheet(COMPACT_INPUT_STYLESHEET)
        center_text_vertically(widget)

    def _make_item_row(self):
        row_index = len(self.item_rows)
        y = COMPACT_ITEM_ROW_TOP + row_index * COMPACT_ITEM_ROW_SPACING
        widgets = []
        for _key, x, width in COMPACT_ITEM_ROW_COLUMNS:
            widget = QTextBrowser(self.centralwidget)
            widget.setGeometry(x, y, width, COMPACT_ITEM_ROW_HEIGHT)
            self._style_row_widget(widget)
            widget.show()
            widgets.append(widget)
        return widgets

    def _rebuild_entry_chain(self):
        fields = [self.Name_textBrowse, self.Address_textBrowse]
        for row in self.item_rows:
            fields.extend(row)

        for field in fields:
            field.setTabChangesFocus(True)
        for current, following in zip(fields, fields[1:]):
            self.setTabOrder(current, following)
            _advance_on_enter(current, following.setFocus)
        self.setTabOrder(fields[-1], self.Go_pushButton)
        _advance_on_enter(fields[-1], self.Go_pushButton.click)

    def _update_bottom_layout(self):
        row_count = len(self.item_rows)
        last_row_bottom = COMPACT_ITEM_ROW_TOP + (row_count - 1) * COMPACT_ITEM_ROW_SPACING + COMPACT_ITEM_ROW_HEIGHT
        needed_top = last_row_bottom + COMPACT_ITEM_ROW_GAP
        delta = max(0, needed_top - self._bottom_cluster_top)

        for widget, origin in self._bottom_cluster_origin.items():
            widget.move(origin.x(), origin.y() + delta)
        self.setFixedSize(self._base_window_size.width(), self._base_window_size.height() + delta)

        self.Plus_pushButton.setEnabled(row_count < MAX_ITEM_ROWS)
        self.Minus_pushButton.setEnabled(row_count > 1)

    def add_item_row(self):
        if len(self.item_rows) >= MAX_ITEM_ROWS:
            return
        widgets = self._make_item_row()
        self.item_rows.append(widgets)
        self._rebuild_entry_chain()
        self._update_bottom_layout()
        widgets[0].setFocus()

    def remove_item_row(self):
        if len(self.item_rows) <= 1:
            return
        widgets = self.item_rows.pop()
        for widget in widgets:
            widget.deleteLater()
        self._rebuild_entry_chain()
        self._update_bottom_layout()

    def open_full_page(self):
        rows = [
            (
                item_w.toPlainText().strip(),
                desc_w.toPlainText().strip(),
                qty_w.toPlainText().strip(),
                price_w.toPlainText().strip(),
            )
            for item_w, desc_w, qty_w, price_w in self.item_rows
        ]
        prefill = {
            "date": self.Date_textBrowser.toPlainText().strip(),
            "number": self.Number_textBrowse.toPlainText().strip(),
            "name": self.Name_textBrowse.toPlainText().strip(),
            "address": self.Address_textBrowse.toPlainText().strip(),
            "rows": rows,
        }
        self._next_window = self.full_page_window_class(prefill=prefill)
        frame = self._next_window.frameGeometry()
        frame.moveCenter(self.frameGeometry().center())
        self._next_window.move(frame.topLeft())
        self._next_window.show()
        self.close()


class Quotation2Window(QMainWindow):
    def __init__(self, prefill=None):
        super().__init__()
        load_ui(QUOTATION2_UI, self)
        self.setWindowTitle("견적서")
        make_window_scrollable(self)

        self.textEdit_6.setText(QDate.currentDate().toString("yyyy-MM-dd"))
        self.textEdit_7.setText(next_document_number("BN-QUT", QUOTATION_SEQ_FILE))

        style_item_form(
            self,
            ITEM_ROWS,
            [self.textEdit_6, self.textEdit_7, self.ClientName_textEdit, self.ClientAddress_textEdit],
        )
        connect_amount_calculations(self, ITEM_ROWS)
        entry_fields = build_entry_chain(
            self, ITEM_ROWS, [self.ClientName_textEdit, self.ClientAddress_textEdit], self.Save_pushButton
        )
        self.textEdit_2.setTabChangesFocus(True)

        if prefill:
            self._apply_prefill(prefill)

        entry_fields[0].setFocus()

        self._saving = False
        self.Save_pushButton.clicked.connect(self.save_quotation)

    def _apply_prefill(self, data):
        self.ClientName_textEdit.setPlainText(data.get("name", ""))
        self.ClientAddress_textEdit.setPlainText(data.get("address", ""))
        for row_names, (item, desc, qty, price) in zip(ITEM_ROWS, data.get("rows", [])):
            item_name, desc_name, qty_name, price_name, _amount_name = row_names
            getattr(self, item_name).setPlainText(item)
            getattr(self, desc_name).setPlainText(desc)
            getattr(self, qty_name).setPlainText(qty)
            getattr(self, price_name).setPlainText(price)

    def save_quotation(self):
        if self._saving:
            return
        self._saving = True
        self.Save_pushButton.setEnabled(False)
        winsound.PlaySound(CLICK_SOUND, winsound.SND_FILENAME | winsound.SND_ASYNC)

        try:
            doc_no = save_filled_document(
                self,
                item_rows=ITEM_ROWS,
                header_cell_map=HEADER_CELL_MAP,
                remarks_cell=REMARKS_CELL,
                remarks_widget_name="textEdit_2",
                xlsx_template=QUOTATION_XLSX,
                save_dir=QUOTATION_SAVE_DIR,
                list_xlsx_path=QUOTATION_LIST_XLSX,
                name_widget_name="ClientName_textEdit",
                date_widget_name="textEdit_6",
                number_widget_name="textEdit_7",
            )
            QMessageBox.information(self, "저장 완료", f"{doc_no} 견적서가 저장되었습니다.")
        except Exception as exc:
            QMessageBox.critical(self, "견적서 저장 실패", f"견적서를 저장하는 중 오류가 발생했습니다.\n\n{exc}")
        finally:
            self._saving = False
            self.Save_pushButton.setEnabled(True)


class QuotationWindow(CompactItemFormWindow):
    ui_path = QUOTATION_UI
    window_title = "견적서"
    number_prefix = "BN-QUT"
    seq_file = QUOTATION_SEQ_FILE
    bottom_cluster_names = QUOTATION_BOTTOM_CLUSTER_WIDGET_NAMES
    full_page_window_class = Quotation2Window


class DeliveryNote2Window(QMainWindow):
    def __init__(self, prefill=None):
        super().__init__()
        load_ui(DELIVERY_NOTE2_UI, self)
        self.setWindowTitle("거래명세서")
        make_window_scrollable(self)

        self.Date_textEdit.setText(QDate.currentDate().toString("yyyy-MM-dd"))
        self.textEdit_7.setText(next_document_number("BN-INV", DELIVERY_NOTE_SEQ_FILE))

        style_item_form(
            self,
            DELIVERY_ITEM_ROWS,
            [self.Date_textEdit, self.textEdit_7, self.ClientNametextEdit, self.ClientAdd_textEdit],
        )
        connect_amount_calculations(self, DELIVERY_ITEM_ROWS)
        entry_fields = build_entry_chain(
            self, DELIVERY_ITEM_ROWS, [self.ClientNametextEdit, self.ClientAdd_textEdit], self.Save_pushButton
        )
        self.textEdit_2.setTabChangesFocus(True)

        if prefill:
            self._apply_prefill(prefill)

        entry_fields[0].setFocus()

        self._saving = False
        self.Save_pushButton.clicked.connect(self.save_delivery_note)

    def _apply_prefill(self, data):
        self.ClientNametextEdit.setPlainText(data.get("name", ""))
        self.ClientAdd_textEdit.setPlainText(data.get("address", ""))
        for row_names, (item, desc, qty, price) in zip(DELIVERY_ITEM_ROWS, data.get("rows", [])):
            item_name, desc_name, qty_name, price_name, _amount_name = row_names
            getattr(self, item_name).setPlainText(item)
            getattr(self, desc_name).setPlainText(desc)
            getattr(self, qty_name).setPlainText(qty)
            getattr(self, price_name).setPlainText(price)

    def save_delivery_note(self):
        if self._saving:
            return
        self._saving = True
        self.Save_pushButton.setEnabled(False)
        winsound.PlaySound(CLICK_SOUND, winsound.SND_FILENAME | winsound.SND_ASYNC)

        try:
            doc_no = save_filled_document(
                self,
                item_rows=DELIVERY_ITEM_ROWS,
                header_cell_map=DELIVERY_HEADER_CELL_MAP,
                remarks_cell=DELIVERY_REMARKS_CELL,
                remarks_widget_name="textEdit_2",
                xlsx_template=DELIVERY_NOTE_XLSX,
                save_dir=DELIVERY_NOTE_SAVE_DIR,
                list_xlsx_path=DELIVERY_NOTE_LIST_XLSX,
                name_widget_name="ClientNametextEdit",
                date_widget_name="Date_textEdit",
                number_widget_name="textEdit_7",
            )
            QMessageBox.information(self, "저장 완료", f"{doc_no} 거래명세서가 저장되었습니다.")
        except Exception as exc:
            QMessageBox.critical(self, "거래명세서 저장 실패", f"거래명세서를 저장하는 중 오류가 발생했습니다.\n\n{exc}")
        finally:
            self._saving = False
            self.Save_pushButton.setEnabled(True)


class DeliveryNoteWindow(CompactItemFormWindow):
    ui_path = DELIVERY_NOTE_UI
    window_title = "거래명세서"
    number_prefix = "BN-INV"
    seq_file = DELIVERY_NOTE_SEQ_FILE
    bottom_cluster_names = DELIVERY_NOTE_BOTTOM_CLUSTER_WIDGET_NAMES
    full_page_window_class = DeliveryNote2Window

    def __init__(self):
        super().__init__()
        self.Go_pushButton_2.clicked.connect(self.load_reference_document)

    def load_reference_document(self):
        path, _filter = QFileDialog.getOpenFileName(
            self,
            "문서 참조 - 불러오기",
            DELIVERY_NOTE_SAVE_DIR,
            "Excel/PDF 파일 (*.xlsx *.pdf)",
        )
        if not path:
            return

        try:
            data = parse_saved_xlsx(path) if path.lower().endswith(".xlsx") else parse_saved_pdf(path)
        except Exception as exc:
            QMessageBox.critical(self, "불러오기 실패", f"문서를 불러오는 중 오류가 발생했습니다.\n\n{exc}")
            return

        self.Name_textBrowse.setPlainText(data["name"])
        self.Address_textBrowse.setPlainText(data["address"])

        rows = data["rows"] or [("", "", "", "")]
        target_count = min(len(rows), MAX_ITEM_ROWS)
        while len(self.item_rows) < target_count:
            self.add_item_row()
        while len(self.item_rows) > max(target_count, 1):
            self.remove_item_row()

        for row_widgets, (item, desc, qty, price) in zip(self.item_rows, rows):
            item_w, desc_w, qty_w, price_w = row_widgets
            item_w.setPlainText(item)
            desc_w.setPlainText(desc)
            qty_w.setPlainText(qty)
            price_w.setPlainText(price)


class DocWindow(QMainWindow):
    """Compact entry form (Doc.ui): fill the header fields here, then Go hands
    everything off to Doc2Window for the full-page body/approval/save."""

    FIELD_NAMES = (
        "Person_textBrowse",  # 담당자
        "_textBrowser",  # 문서번호
        "Number_textBrowse",  # 시행일자
        "Name_textBrowse",  # 경유
        "Address_textBrowse",  # 수신
        "Item_textBrowse",  # 제목
    )

    def __init__(self):
        super().__init__()
        load_ui(DOC_UI, self)
        self.setFixedSize(self.size())
        self.setWindowTitle("공문서")

        # 문서번호/시행일자 are auto-filled below and left read-only; the rest are typed in.
        auto_filled = ("_textBrowser", "Number_textBrowse")
        for name in self.FIELD_NAMES:
            widget = getattr(self, name)
            widget.setReadOnly(name in auto_filled)
            widget.setLineWrapMode(QTextEdit.NoWrap)
            widget.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            widget.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            widget.setStyleSheet(COMPACT_INPUT_STYLESHEET)
            font = widget.font()
            font.setPointSize(12)
            widget.setFont(font)
            center_text_vertically(widget)

        self._textBrowser.setText(next_document_number("BN-DLM", DOC_SEQ_FILE))
        self.Number_textBrowse.setText(QDate.currentDate().toString("yyyy-MM-dd"))

        entry_fields = [getattr(self, name) for name in self.FIELD_NAMES]
        for field in entry_fields:
            field.setTabChangesFocus(True)
        for current, following in zip(entry_fields, entry_fields[1:]):
            self.setTabOrder(current, following)
            _advance_on_enter(current, following.setFocus)
        self.setTabOrder(entry_fields[-1], self.Go_pushButton)
        _advance_on_enter(entry_fields[-1], self.Go_pushButton.click)
        entry_fields[0].setFocus()

        self._loaded_body = None
        self._next_window = None
        self.Go_pushButton.clicked.connect(self.open_doc2)
        self.Back_pushButton.clicked.connect(self.close)
        self.Go_pushButton_2.clicked.connect(self.load_reference_document)

    def open_doc2(self):
        prefill = {
            "number": self._textBrowser.toPlainText().strip(),
            "date": self.Number_textBrowse.toPlainText().strip(),
            "person": self.Person_textBrowse.toPlainText().strip(),
            "routing": self.Name_textBrowse.toPlainText().strip(),
            "recipient": self.Address_textBrowse.toPlainText().strip(),
            "title": self.Item_textBrowse.toPlainText().strip(),
            "body": self._loaded_body,
        }
        self._next_window = Doc2Window(prefill=prefill)
        frame = self._next_window.frameGeometry()
        frame.moveCenter(self.frameGeometry().center())
        self._next_window.move(frame.topLeft())
        self._next_window.show()
        self.close()

    def load_reference_document(self):
        path, _filter = QFileDialog.getOpenFileName(
            self, "문서 참조 - 불러오기", DOC_SAVE_DIR, "Excel/PDF 파일 (*.xlsx *.pdf)"
        )
        if not path:
            return

        try:
            data = parse_saved_doc_xlsx(path) if path.lower().endswith(".xlsx") else parse_saved_doc_pdf(path)
        except Exception as exc:
            QMessageBox.critical(self, "불러오기 실패", f"문서를 불러오는 중 오류가 발생했습니다.\n\n{exc}")
            return

        self.Person_textBrowse.setPlainText(data["person"])
        self.Name_textBrowse.setPlainText(data["routing"])
        self.Address_textBrowse.setPlainText(data["recipient"])
        self.Item_textBrowse.setPlainText(data["title"])
        self._loaded_body = data["body"]


class Doc2Window(QMainWindow):
    def __init__(self, prefill=None):
        super().__init__()
        load_ui(DOC2_UI, self)
        self.setWindowTitle("공문서")
        make_window_scrollable(self)

        for text_edit in self.centralwidget.findChildren(QTextEdit):
            center_text_vertically(text_edit)

        single_line_widgets = [
            self.Person_textEdit,
            self.Number_textEdit,
            self.Date_textEdit,
            self.textEdit_3,
            self.textEdit_4,
            self.Title_textEdit,
            self.textEdit_6,
            self.textEdit_7,
        ]
        for widget in single_line_widgets:
            widget.setLineWrapMode(QTextEdit.NoWrap)
            widget.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        self.Number_textEdit.setText(
            prefill["number"] if prefill and prefill.get("number") else next_document_number("BN-DLM", DOC_SEQ_FILE)
        )
        self.Date_textEdit.setText(
            prefill["date"] if prefill and prefill.get("date") else QDate.currentDate().toString("yyyy-MM-dd")
        )

        # The footer "담당자" line is the same person as the letterhead one - shown
        # read-only here and kept in sync, so it's only ever typed once.
        self.textEdit_6.setReadOnly(True)
        self.Person_textEdit.textChanged.connect(
            lambda: self.textEdit_6.setPlainText(self.Person_textEdit.toPlainText())
        )

        entry_fields = [
            self.Person_textEdit,
            self.textEdit_3,
            self.textEdit_4,
            self.Title_textEdit,
            self.MainBody_textEdit,
            self.textEdit_7,
        ]
        for field in entry_fields:
            field.setTabChangesFocus(True)
        for current, following in zip(entry_fields, entry_fields[1:]):
            self.setTabOrder(current, following)
            if current is not self.MainBody_textEdit:
                _advance_on_enter(current, following.setFocus)
        self.setTabOrder(entry_fields[-1], self.Save_pushButton)
        _advance_on_enter(entry_fields[-1], self.Save_pushButton.click)
        # MainBody_textEdit is free-form multi-line text, so Enter there is left alone
        # to insert a newline instead of jumping focus.

        if prefill:
            self.Person_textEdit.setPlainText(prefill.get("person", ""))
            self.textEdit_3.setPlainText(prefill.get("routing", ""))
            self.textEdit_4.setPlainText(prefill.get("recipient", ""))
            self.Title_textEdit.setPlainText(prefill.get("title", ""))
            if prefill.get("body"):
                self.MainBody_textEdit.setPlainText(prefill["body"])

        for button_name, _cell, _role in DOC_APPROVAL_STAGES:
            button = getattr(self, button_name)
            button.clicked.connect(lambda _checked=False, b=button: self._cycle_approval(b))

        entry_fields[0].setFocus()

        self._saving = False
        self.Save_pushButton.clicked.connect(self.save_doc)

    def _cycle_approval(self, button):
        """결재대기 -> 승인 -> 반려 -> 결재대기 ... one click per state change."""
        state = button.property("approval_state") or "pending"
        next_state = {"pending": "approved", "approved": "rejected", "rejected": "pending"}[state]
        button.setProperty("approval_state", next_state)

        today = QDate.currentDate().toString("yyyy-MM-dd")
        if next_state == "approved":
            button.setIcon(QIcon(DOC_APPROVE_ICON))
            button.setText(f"승인\n{today}")
            button.setStyleSheet("background-color: rgb(198, 239, 206); color: rgb(30, 90, 30);")
        elif next_state == "rejected":
            button.setIcon(QIcon(DOC_REJECT_ICON))
            button.setText(f"반려\n{today}")
            button.setStyleSheet("background-color: rgb(245, 205, 203); color: rgb(120, 30, 30);")
        else:
            button.setIcon(QIcon())
            button.setText("결재대기")
            button.setStyleSheet("color: rgb(209, 209, 209);")

    def save_doc(self):
        if self._saving:
            return
        self._saving = True
        self.Save_pushButton.setEnabled(False)
        winsound.PlaySound(CLICK_SOUND, winsound.SND_FILENAME | winsound.SND_ASYNC)

        try:
            workbook = openpyxl.load_workbook(DOC_XLSX)
            sheet = workbook.active

            for cell, widget_name in DOC_HEADER_CELL_MAP:
                sheet[cell] = getattr(self, widget_name).toPlainText().strip()

            person = self.Person_textEdit.toPlainText().strip()
            sheet[DOC_PERSON_CELL] = f"{DOC_PERSON_LABEL}     {person}" if person else DOC_PERSON_LABEL
            sheet[DOC_FOOTER_PERSON_CELL] = f"담당자   :   {person}"
            coop = self.textEdit_7.toPlainText().strip()
            sheet[DOC_FOOTER_COOP_CELL] = f"협조자   :   {coop}"

            body_lines = self.MainBody_textEdit.toPlainText().split("\n")
            for row, line in zip(DOC_BODY_ROWS, body_lines):
                sheet[f"A{row}"] = line

            for button_name, cell, _role in DOC_APPROVAL_STAGES:
                button = getattr(self, button_name)
                if (button.property("approval_state") or "pending") != "pending":
                    sheet[cell] = button.text()

            doc_no = self.Number_textEdit.toPlainText().strip()
            export_document_files(workbook, doc_no, DOC_SAVE_DIR, self.centralwidget)

            list_date = QDate.fromString(self.Date_textEdit.toPlainText().strip(), "yyyy-MM-dd").toString("yyMMdd")
            append_to_doc_list(
                DOC_LIST_XLSX,
                list_date,
                person,
                doc_no,
                self.textEdit_3.toPlainText().strip(),
                self.textEdit_4.toPlainText().strip(),
                self.Title_textEdit.toPlainText().strip(),
            )

            QMessageBox.information(self, "저장 완료", f"{doc_no} 공문서가 저장되었습니다.")
        except Exception as exc:
            QMessageBox.critical(self, "공문서 저장 실패", f"공문서를 저장하는 중 오류가 발생했습니다.\n\n{exc}")
        finally:
            self._saving = False
            self.Save_pushButton.setEnabled(True)


class HomeWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        load_ui(HOME_UI, self)
        self.setFixedSize(self.size())
        self.setWindowTitle("Benchsoft")
        self._children = []

        # The caption labels are laid out on top of the icon buttons, so by default they
        # swallow every click that lands on the text and the button never fires.
        for label in self.findChildren(QLabel):
            label.setAttribute(Qt.WA_TransparentForMouseEvents, True)

        self.pushButton.clicked.connect(self.open_doc)
        self.pushButton_4.clicked.connect(self.open_doc_list)
        self.pushButton_2.clicked.connect(self.open_quotation)
        self.pushButton_5.clicked.connect(self.open_quotation_list)
        self.pushButton_3.clicked.connect(self.open_delivery_note)
        self.pushButton_6.clicked.connect(self.open_delivery_note_list)

    def _open_centered(self, child):
        frame = child.frameGeometry()
        frame.moveCenter(self.frameGeometry().center())
        child.move(frame.topLeft())
        # Keep a reference so the window is not garbage collected, and drop the ones the
        # user has already closed instead of holding them for the whole session.
        self._children = [c for c in self._children if c.isVisible()]
        self._children.append(child)
        child.show()
        child.raise_()
        child.activateWindow()

    def open_quotation(self):
        self._open_centered(QuotationWindow())

    def open_quotation_list(self):
        os.startfile(QUOTATION_LIST_XLSX)

    def open_delivery_note(self):
        self._open_centered(DeliveryNoteWindow())

    def open_delivery_note_list(self):
        os.startfile(DELIVERY_NOTE_LIST_XLSX)

    def open_doc(self):
        self._open_centered(DocWindow())

    def open_doc_list(self):
        os.startfile(DOC_LIST_XLSX)


def main():
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    app = QApplication(sys.argv)
    window = HomeWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
